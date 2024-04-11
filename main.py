from bs4 import BeautifulSoup
import requests
import openpyxl as xl
from openpyxl.utils.exceptions import IllegalCharacterError
import re


class Excel_Con():
    def __init__(self):
        self.wb_jp = xl.Workbook()
        self.ws_jp = self.wb_jp.active
        """self.dict_en = {
            'Company Name': '会社名',
            'Representative Person': '代表者',
            'Business Description': '事業内容',
            'Product Line': '取扱品目',
            'Address': '住所',
            'Website': 'ウェブサイト',
            'Establish Date': '設立日',
            'Capital': '資本金',
            'Parent Company': '親会社',
            'Shareholder': '株主',
            'Employee': '従業員数',
            'Account Period': '決算期',
            'Bank': '取引銀行',
            'BOI': 'BOI',
            'Certificate': '証明書',
            'Major Customer': '主要取引先'
        }

        self.dict_th = {
            'ชื่อบริษัท': '会社名',
            'ผู้รับผิดชอบ': '代表者',
            'รายละเอียดธุรกิจ': '事業内容',
            'สินค้า / บริการ': '取扱品目',
            'ที่อยู่': '住所',
            'เว็บไซต์': 'ウェブサイト',
            'วันที่ก่อตั้ง': '設立日',
            'ทุนจดทะเบียน': '資本金',
            'บริษัทแม่': '親会社',
            'ผู้ถือหุ้น': '株主',
            'จำนวนพนักงาน': '従業員数',
            'รอบตัดบัญชี': '決算期',
            'ธนาคาร': '取引銀行',
            'BOI': 'BOI',
            'ใบรับรอง': '証明書',
            'ลูกค้าสำคัญ': '主要取引先'
        }

        self.titles = [
            'No.',
            '会社名',
            '代表者',
            '事業内容',
            '取扱品目',
            '住所',
            'ウェブサイト',
            '設立日',
            '資本金',
            '親会社',
            '株主',
            '従業員数',
            '決算期',
            '取引銀行',
            'BOI',
            '証明書',
            '主要取引先'
        ]"""
        # self.ws.append(self.titles)
        self.ws_jp.cell(1, 1).value = 'Index'
        self.ws_jp.cell(1, 2).value = '業種'
        self.wb_jp.save(filename='output_jp.xlsx')

        self.wb_en = xl.Workbook()
        self.ws_en = self.wb_en.active
        self.ws_en.cell(1, 1).value = 'Index'
        self.ws_en.cell(1, 2).value = 'Category'
        self.wb_en.save(filename='output_en.xlsx')

        self.wb_th = xl.Workbook()
        self.ws_th = self.wb_th.active
        self.ws_th.cell(1, 1).value = 'Index'
        self.ws_th.cell(1, 2).value = 'หมวดธุรกิจ'
        self.wb_th.save(filename='output_th.xlsx')

        self.i_jp = 1
        self.i_en = 1
        self.i_th = 1

    def writing_excel(self, category_result):
        for each_result in category_result:
            print(each_result)
            if each_result['lang'] == 'jp':
                self.i_jp += 1
                self.ws_jp.cell(self.i_jp, 1).value = self.i_jp - 1
                self.ws_jp.cell(self.i_jp, 2).value = each_result['category']
                for each_key in each_result:
                    if each_key == 'lang' or each_key == 'category':
                        continue
                    for j in range(3, 101):
                        if self.ws_jp.cell(1, j).value:
                            if str(each_key).strip() == self.ws_jp.cell(1, j).value:
                                try:
                                    self.ws_jp.cell(self.i_jp, j).value = each_result[each_key]
                                except IllegalCharacterError:
                                    pass
                                break
                        else:
                            self.ws_jp.cell(1, j).value = str(each_key).strip()
                            try:
                                self.ws_jp.cell(self.i_jp, j).value = each_result[each_key]
                            except IllegalCharacterError:
                                pass
                            break
                self.wb_jp.save(filename='output_jp.xlsx')

            elif each_result['lang'] == 'en':
                self.i_en += 1
                self.ws_en.cell(self.i_en, 1).value = self.i_en - 1
                self.ws_en.cell(self.i_en, 2).value = each_result['category']
                for each_key in each_result:
                    if each_key == 'lang' or each_key == 'category':
                        continue
                    for j in range(3, 101):
                        if self.ws_en.cell(1, j).value:
                            if str(each_key).strip() == self.ws_en.cell(1, j).value:
                                try:
                                    self.ws_en.cell(self.i_en, j).value = each_result[each_key]
                                except IllegalCharacterError:
                                    pass
                                break
                        else:
                            self.ws_en.cell(1, j).value = str(each_key).strip()
                            try:
                                self.ws_en.cell(self.i_en, j).value = each_result[each_key]
                            except IllegalCharacterError:
                                pass
                            break
                self.wb_en.save(filename='output_en.xlsx')

            else:
                self.i_th += 1
                self.ws_th.cell(self.i_th, 1).value = self.i_th - 1
                self.ws_th.cell(self.i_th, 2).value = each_result['category']
                for each_key in each_result:
                    if each_key == 'lang' or each_key == 'category':
                        continue
                    for j in range(3, 101):
                        if self.ws_th.cell(1, j).value:
                            if str(each_key).strip() == self.ws_th.cell(1, j).value:
                                try:
                                    self.ws_th.cell(self.i_th, j).value = each_result[each_key]
                                except IllegalCharacterError:
                                    pass
                                break
                        else:
                            self.ws_th.cell(1, j).value = str(each_key).strip()
                            try:
                                self.ws_th.cell(self.i_th, j).value = each_result[each_key]
                            except IllegalCharacterError:
                                pass
                            break
                self.wb_th.save(filename='output_th.xlsx')

    def close_excel(self):
        self.wb_jp.save(filename='output_jp.xlsx')
        self.wb_jp.close()

        self.wb_en.save(filename='output_en.xlsx')
        self.wb_en.close()

        self.wb_th.save(filename='output_th.xlsx')
        self.wb_th.close()


class Scraping():
    def __init__(self):
        self.url_root = 'https://www.fact-link.com/'
        self.headers = {
            'authority': 'www.fact-link.com',
            'method': 'GET',
            'path': "/",
            'scheme': 'https',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': "ja,zh-TH;q=0.9,zh;q=0.8,th-TH;q=0.7,th;q=0.6,en-TH;q=0.5,en-US;q=0.4,en;q=0.3",
            'Cache-Control': 'max-age=0',
            'Cookie': 'lang=jp; __utmz=118580510.1711683656.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); _ga=GA1.1.827601414.1711683657; cookieControl=true; cookieControlPrefs=["preferences","analytics","marketing"]; PHPSESSID=poa9f0pcjql3tqcra407laanr2; __utma=118580510.57685764.1711683656.1712107907.1712121888.5; __utmc=118580510; _ga_T8H0BLR4NB=GS1.1.1712121769.5.1.1712121891.0.0.0; __utmt=1; __utmb=118580510.7.10.1712121888; _ga_5SYSWMEFJ3=GS1.1.1712121887.4.1.1712122795.0.0.0',
            'Referer': 'https://www.fact-link.com/',
            'Sec-Ch-Ua': '"Google Chrome";v="123", "Not:A-Brand";v="8", "Chromium";v="123"',
            'Sec-Ch-Ua-Mobile': '?0',
            'Sec-Ch-Ua-Platform': '"Windows"',
            'Sec-Fetch-Dest': 'document',
            'Sec-Fetch-Mode': 'navigate',
            'Sec-Fetch-Site': 'same-origin',
            'Upgrade-Insecure-Requests': '1',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36'
        }
        self.first_category_url = ''

    def reading_category_links(self):
        rs = requests.get(url=self.url_root, headers=self.headers)
        soup = BeautifulSoup(rs.text, features='html.parser')
        section = soup.find('section', class_='category')

        menu_sp = BeautifulSoup(str(section), features='html.parser')
        menu = menu_sp.find('div', class_='menu')

        category_sp = BeautifulSoup(str(menu), features='html.parser')
        category = category_sp.find_all('a')

        data = {}
        for each in category:
            data[each.text] = self.url_root + each['href']

        return data

    def reading_company_links(self, category_name, category_url, data):
        category_name = category_name

        rs = requests.get(url=category_url, headers=self.headers)
        soup = BeautifulSoup(rs.text, features='html.parser')
        divs_cp = soup.find_all('div', class_='cp')

        next_btn = soup.find('a', text='次へ')

        data = data
        print(data, '!!!!!!!!!!!!!!!')
        for div_cp in divs_cp:
            div_cp_sp = BeautifulSoup(str(div_cp), features='html.parser')

            p_corp = div_cp_sp.find('p', class_='corp')
            company_name = p_corp.text

            a_jp = div_cp_sp.find('a', class_='jp')
            if a_jp != None:
                company_link = a_jp['href']
                lang = 'jp'
            else:
                a_en = div_cp_sp.find('a', class_='en')
                if a_en != None:
                    company_link = a_en['href']
                    lang = 'en'
                else:
                    a_th = div_cp_sp.find('a', class_='th')
                    if a_th != None:
                        company_link = a_th['href']
                        lang = 'th'
                    else:
                        company_link = None
                        lang = None

            if company_link:
                if 'mem_content' in company_link:
                    company_link = self.get_profile_link(content_link=self.url_root + str(company_link), lang=lang)
                else:
                    company_link = self.url_root + str(company_link)

                if company_link:
                    company_info = self.get_company_details(company_link=company_link, lang=lang,
                                                            category_name=category_name)
                    data.append(company_info)

            print(company_name)
            print(company_link)

        if next_btn:
            # print(next_btn, '\n', next_btn['href'])
            offset_str = str(next_btn['href']).replace('?', '&')
            return self.reading_company_links(category_name=category_name,
                                              category_url=self.first_category_url + offset_str, data=data)
        else:
            return data

    def get_profile_link(self, content_link, lang):
        rs = requests.get(url=content_link, headers=self.headers)
        soup = BeautifulSoup(rs.text, features='html.parser')
        link_a = soup.find('a', href=re.compile(rf'mem_profile\.php\?pl={lang}'))
        try:
            link = link_a['href']
        except:
            return None
        if 'www.fact-link.com' in link:
            return link
        else:
            return self.url_root + link

    def get_company_details(self, company_link, lang, category_name):
        company_info = {'lang': lang, 'category': category_name}
        rs = requests.get(url=company_link, headers=self.headers)
        # print(rs.text)
        soup = BeautifulSoup(rs.text, features='html.parser')
        profile_sets = soup.find_all('div', class_='profileset')
        for profile_set in profile_sets:
            profile_set_sp = BeautifulSoup(str(profile_set), features='html.parser')
            profile_label = profile_set_sp.find('div', class_='profilelabel')
            profile_detail = profile_set_sp.find('div', class_='profiledetail')
            # print(profile_label)
            # print(profile_detail)
            try:
                header = profile_label.text
            except AttributeError:
                continue
            text = profile_detail.text
            company_info[header] = text
            # print(header, '--------', text)
        return company_info


if __name__ == '__main__':
    SC = Scraping()
    data = SC.reading_category_links()

    EC = Excel_Con()

    for key in data:
        # print(key, data[key])
        category_name = key
        category_url = data[key]
        SC.first_category_url = category_url
        category_result = SC.reading_company_links(category_name=category_name, category_url=category_url, data=[])
        print(category_result)
        EC.writing_excel(category_result=category_result)

    EC.close_excel()
