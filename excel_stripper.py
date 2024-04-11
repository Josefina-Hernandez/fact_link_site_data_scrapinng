import openpyxl as xl

class Excel_Con():
    def __init__(self):
        self.wb = xl.load_workbook(filename='output_th(Final_Revised).xlsx')
        self.ws = self.wb.active

    def running(self):
        for i in range(1, self.ws.max_row+1):
            for j in range(1, self.ws.max_column+1):
                print(f'Scanning ({i, j})......')
                if self.ws.cell(row=i, column=j).value:
                    self.ws.cell(row=i, column=j).value = str(self.ws.cell(row=i, column=j).value).strip()

        self.wb.save(filename='output_th(Final_Revised).xlsx')
        self.wb.close()
        print('Finished!')

if __name__ == '__main__':
    EC = Excel_Con()
    EC.running()