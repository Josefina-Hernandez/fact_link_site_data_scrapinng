import openpyxl as xl
import re

wb = xl.load_workbook('./input_for_2/output_en(Final_Revised).xlsx')
ws = wb.active

def if_mail(str):
    if '@' not in str:
        return False
    result = []
    pattern = re.compile(r' |\xa0|\u3000|\n|：|※')
    splited_words = pattern.split(str)
    for each_word in splited_words:
        if '@' in each_word:
            result.append(each_word)

    return result

def reading_data():
    data = []
    for i in range(2, ws.max_row+1):
        line_mail = []
        for j in range(1, ws.max_column+1):
            cell_value = ws.cell(row=i, column=j).value
            if not cell_value:
                continue
            cell_mail = if_mail(str=cell_value)
            if cell_mail:
                line_mail=cell_mail
        #print(i, line_mail)
        if len(line_mail) > 1:
            str = ''
            for each in line_mail:
                str += each + '\n'
            str = str[:-1]
        else:
            if line_mail:
                str = line_mail[0]
            else:
                str = None
        print(i, str)
        data.append(str)

    wb.close()
    return data

def create_new_excel(data):
    wb = xl.load_workbook(filename='./input_for_2/output_en(Final_Revised).xlsx')
    ws = wb.active

    ws.insert_cols(4)

    wb.save('./output_for_2/output_en(Final_Revised)2.xlsx')
    ws.cell(row=1, column=4).value = 'Mail Address'

    for i in range(2, ws.max_row+1):
        ws.cell(row=i, column=4).value = data[i-2]
    wb.save('./output_for_2/output_en(Final_Revised)2.xlsx')
    wb.close()

if __name__ == '__main__':
    mail_data = reading_data()
    create_new_excel(data=mail_data)